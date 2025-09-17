from io import BytesIO
from decimal import Decimal
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile

from excelread.services.reader import ExcelImporter, UnsupportedFileError
from excelread.models import RABItem

class ExcelReaderTests(TestCase):
    def _xlsx_file(self):
        from openpyxl import Workbook
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append([1, "Pondasi", 10.5, "m3"])
        ws.append([2, "Beton Kolom", "1.000,5", "m3"])  # Indonesian number style
        wb.save(bio)
        return SimpleUploadedFile(
            "rab.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _xls_file(self):
        import xlwt
        bio = BytesIO()
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet1')
        rows = [
            ["No", "Uraian Pekerjaan", "Volume", "Satuan"],
            [1, "Pondasi", 3.25, "m3"],
            ["1.2", "Urugan", "2.500,00", "m3"],
        ]
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
        wb.save(bio)
        return SimpleUploadedFile(
            "rab.xls",
            bio.getvalue(),
            content_type="application/vnd.ms-excel",
        )

    def test_read_xlsx_into_db(self):
        from excelread.services.reader import ExcelImporter
        f = self._xlsx_file()
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 2)
        self.assertEqual(RABItem.objects.count(), 2)
        first = RABItem.objects.order_by("row_index").first()
        self.assertEqual(first.description, "Pondasi")
        self.assertEqual(first.unit, "m3")
        self.assertEqual(first.volume, Decimal("10.5"))

        # Indonesian-style number parsed
        second = RABItem.objects.order_by("row_index")[1]
        self.assertEqual(second.volume, Decimal("1000.5"))

    def test_read_xls_into_db(self):
        from excelread.services.reader import ExcelImporter
        f = self._xls_file()
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 2)
        self.assertEqual(RABItem.objects.count(), 2)

    def test_header_variants(self):
        # Accept e.g. "Uraian", "Deskripsi", "Qty" for robustness
        from openpyxl import Workbook
        from excelread.services.reader import ExcelImporter

        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian", "Qty", "Satuan"])
        ws.append(["A", "Pekerjaan A", "12", "m2"])
        wb.save(bio)
        f = SimpleUploadedFile(
            "alt.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 1)
        item = RABItem.objects.get()
        self.assertEqual(item.number, "A")
        self.assertEqual(item.unit, "m2")
        self.assertEqual(item.volume, Decimal("12"))

    def test_reject_unsupported_type(self):
        from excelread.services.reader import ExcelImporter, UnsupportedFileError
        bad = SimpleUploadedFile("data.csv", b"a,b,c\n", content_type="text/csv")
        with self.assertRaises(UnsupportedFileError):
            ExcelImporter().import_file(bad)
