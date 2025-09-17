from django.core.files.uploadedfile import UploadedFile
from decimal import Decimal
from openpyxl import load_workbook
import xlrd
from excelread.models import RABItem

class UnsupportedFileError(Exception):
    pass

class ExcelImporter:
    def import_file(self, file: UploadedFile) -> int:
        name = (file.name or "").lower()
        if name.endswith(".xlsx"):
            rows = self._read_xlsx(file)
        elif name.endswith(".xls"):
            rows = self._read_xls(file)
        else:
            raise UnsupportedFileError("Only .xls and .xlsx are supported")

        count = 0
        for idx, row in enumerate(rows, start=1):
            # Skip header row
            if idx == 1:
                continue
            if not row or all(v in (None, "", " ") for v in row):
                continue

            number, description, volume, unit = row[:4]

            RABItem.objects.create(
                number=str(number).strip(),
                description=str(description).strip(),
                volume=self._to_decimal(volume),
                unit=str(unit).strip(),
                source_filename=file.name,
                row_index=idx,
            )
            count += 1
        return count

    def _read_xlsx(self, file: UploadedFile):
        file.seek(0)
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def _read_xls(self, file: UploadedFile):
        file.seek(0)
        wb = xlrd.open_workbook(file_contents=file.read())
        sh = wb.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]

    def _to_decimal(self, value):
        try:
            return Decimal(str(value).replace(",", "").strip())
        except Exception:
            return Decimal("0")
