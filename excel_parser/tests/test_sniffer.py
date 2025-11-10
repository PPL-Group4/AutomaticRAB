import tempfile
from django.core.exceptions import ValidationError
from django.test import TestCase
from excel_parser.services.services import ExcelSniffer

try:
    import xlrd
    from xlwt import Workbook as XlsWorkbook
except ImportError:
    xlrd = None
    XlsWorkbook = None

from openpyxl import Workbook


class ExcelSnifferTests(TestCase):
    def _temp_path(self, suffix):
        f = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        f.close()
        return f.name

    def test_valid_xlsx_file(self):
        file_path = self._temp_path(".xlsx")
        wb = Workbook()
        wb.save(file_path)
        sniffer = ExcelSniffer()
        self.assertIsNone(sniffer.is_valid(file_path))

    def test_valid_empty_xlsx_file(self):
        file_path = self._temp_path(".xlsx")
        wb = Workbook()
        wb.save(file_path)
        sniffer = ExcelSniffer()
        self.assertIsNone(sniffer.is_valid(file_path))

    def test_invalid_xlsx_file(self):
        file_path = self._temp_path(".xlsx")
        with open(file_path, "wb") as f:
            f.write(b"not really excel")
        sniffer = ExcelSniffer()
        with self.assertRaises(ValidationError):
            sniffer.is_valid(file_path)

    def test_fake_xlsx_with_wrong_content(self):
        file_path = self._temp_path(".xlsx")
        with open(file_path, "w") as f:
            f.write("Hello world")
        sniffer = ExcelSniffer()
        with self.assertRaises(ValidationError):
            sniffer.is_valid(file_path)

    def test_valid_xls_file(self):
        if not XlsWorkbook:
            self.skipTest("xlrd/xlwt not installed")
        file_path = self._temp_path(".xls")
        wb = XlsWorkbook()
        sheet = wb.add_sheet("Sheet1")
        sheet.write(0, 0, "Hello XLS")
        wb.save(file_path)
        sniffer = ExcelSniffer()
        self.assertIsNone(sniffer.is_valid(file_path))

    def test_invalid_xls_file(self):
        if not xlrd:
            self.skipTest("xlrd not installed")
        file_path = self._temp_path(".xls")
        with open(file_path, "wb") as f:
            f.write(b"not really an xls")
        sniffer = ExcelSniffer()
        with self.assertRaises(ValidationError):
            sniffer.is_valid(file_path)
