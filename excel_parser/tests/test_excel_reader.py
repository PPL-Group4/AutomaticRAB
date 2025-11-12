from io import BytesIO
from decimal import Decimal
from datetime import date
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook
from excel_parser.services.reader import ExcelImporter, UnsupportedFileError
from excel_parser.services.header_mapper import map_headers, find_header_row
from excel_parser.services import create_rab_parser
from excel_parser.models import Project, RabEntry


try:
    from xlwt import Workbook as XlsWorkbook
except ImportError:
    XlsWorkbook = None


class ExcelReaderTests(TestCase):
    def _xlsx_file(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append([1, "Pondasi", 10.5, "m3"])
        ws.append([2, "Beton Kolom", "1.000,5", "m3"])  # Indonesian number style
        wb.save(bio)
        return SimpleUploadedFile(
            "rab.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _xls_file(self):
        if not XlsWorkbook:
            self.skipTest("xlwt not installed")
        bio = BytesIO()
        wb = XlsWorkbook()
        ws = wb.add_sheet("Sheet1")
        rows = [
            ["No", "Uraian Pekerjaan", "Volume", "Satuan"],
            [1, "Pondasi", 3.25, "m3"],
            ["1.2", "Urugan", "2.500,00", "m3"],
        ]
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
        wb.save(bio)
        return SimpleUploadedFile(
            "rab.xls",
            bio.getvalue(),
            content_type="application/vnd.ms-excel",
        )

    def test_read_xlsx_into_db(self):
        f = self._xlsx_file()
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 2)
        self.assertEqual(RabEntry.objects.count(), 2)

        first = RabEntry.objects.order_by("row_index").first()
        self.assertEqual(first.description, "Pondasi")
        self.assertEqual(first.unit, "m3")
        self.assertEqual(first.volume, Decimal("10.5"))

        second = RabEntry.objects.order_by("row_index")[1]
        self.assertEqual(second.volume, Decimal("1000.5"))

    def test_read_xls_into_db(self):
        f = self._xls_file()
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 2)
        self.assertEqual(RabEntry.objects.count(), 2)

    def test_header_variants(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian", "Qty", "Satuan"])
        ws.append(["A", "Pekerjaan A", "12", "m2"])
        wb.save(bio)
        f = SimpleUploadedFile(
            "alt.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 1)
        item = RabEntry.objects.get()
        self.assertEqual(item.item_number, "A")
        self.assertEqual(item.unit, "m2")
        self.assertEqual(item.volume, Decimal("12"))

    def test_reject_unsupported_type(self):
        bad = SimpleUploadedFile("data.csv", b"a,b,c\n", content_type="text/csv")
        with self.assertRaises(UnsupportedFileError):
            ExcelImporter().import_file(bad)


class HeaderMapperTests(TestCase):
    def test_maps_clean_headers(self):
        headers = ["No", "Uraian Pekerjaan", "Volume", "Satuan"]
        mapping, missing, originals = map_headers(headers)
        self.assertIn("no", mapping)
        self.assertIn("uraian", mapping)
        self.assertIn("volume", mapping)
        self.assertIn("satuan", mapping)
        self.assertEqual(missing, [])
        self.assertEqual(originals["uraian"], "Uraian Pekerjaan")

    def test_maps_with_punctuation_and_case(self):
        headers = ["No.", "URAIAN", "VOL.", "satuan "]
        mapping, missing, _ = map_headers(headers)
        self.assertTrue(set(["no", "uraian", "volume", "satuan"]).issubset(mapping.keys()))
        self.assertEqual(missing, [])

    def test_reports_missing_required(self):
        headers = ["No", "Deskripsi"]
        _, missing, _ = map_headers(headers)
        self.assertIn("volume", missing)
        self.assertIn("satuan", missing)

    def test_find_header_row_skips_title_block(self):
        rows = [
            ["PEKERJAAN", None, ":", "PEMBANGUNAN"],
            [None, None, None],
            ["No.", "URAIAN PEKERJAAN", "VOL.", "SATUAN"],
            ["1", "Bata Merah", "10", "m"],
        ]
        idx = find_header_row(rows, scan_first=10)
        self.assertEqual(idx, 2)
