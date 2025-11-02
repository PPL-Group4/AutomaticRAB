from io import BytesIO
from decimal import Decimal
import tempfile
from unittest.mock import patch
from django.core.exceptions import ValidationError
from automatic_job_matching import views
from excel_parser.services.header_mapper import map_headers, find_header_row
from excel_parser.services.validators import validate_excel_file
from excel_parser.services.services import ExcelSniffer
from excel_parser.services import create_rab_parser
from excel_parser.services.reader import ExcelImporter, UnsupportedFileError
from excel_parser.services import reader as reader_mod
from datetime import date
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.sessions.middleware import SessionMiddleware
from django.test import RequestFactory, TestCase, Client, override_settings
from openpyxl import Workbook
from rest_framework.test import APITestCase
from excel_parser.services.reader import preview_file
from django.apps import apps
from django.db import connection, models
from django.test import TransactionTestCase
from excel_parser.models import Project, RabEntry
from excel_parser.services.reader import ExcelImporter
from rencanakan_core.models import RabItem, Rab
from excel_parser import views


SQLITE_DB_SETTINGS = {
    "default": {
        "ENGINE": "django.db.backends.sqlite3",
        "NAME": ":memory:",
    }
}

try:
    import xlrd
    from xlwt import Workbook as XlsWorkbook
except ImportError:
    xlrd = None
    XlsWorkbook = None


def _attach_session(req):
    mw = SessionMiddleware()
    mw.process_request(req)
    req.session.save()
    return req


def _make_xlsx_bytes():
    bio = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan", "Harga Satuan", "Jumlah Harga"])
    ws.append(["1", "Pekerjaan A", "1", "Ls", "1000", "1000"])
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


class FileValidatorTests(TestCase):
    def test_accepts_xlsx_file(self):
        file = SimpleUploadedFile(
            "dummy.xlsx",
            b"dummy",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        try:
            validate_excel_file(file)
        except ValidationError:
            self.fail("Valid .xlsx file was rejected")

    def test_accepts_xls_file(self):
        file = SimpleUploadedFile(
            "dummy.xls",
            b"dummy",
            content_type="application/vnd.ms-excel",
        )
        try:
            validate_excel_file(file)
        except ValidationError:
            self.fail("Valid .xls file was rejected")

    def test_rejects_pdf_file(self):
        file = SimpleUploadedFile("dummy.pdf", b"%PDF", content_type="application/pdf")
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_txt_file(self):
        file = SimpleUploadedFile("dummy.txt", b"Hello", content_type="text/plain")
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_wrong_mimetype_for_excel(self):
        file = SimpleUploadedFile("dummy.xlsx", b"fake excel", content_type="text/plain")
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_empty_file_with_valid_mimetype(self):
        file = SimpleUploadedFile(
            "empty.xlsx",
            b"",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        try:
            validate_excel_file(file)
        except ValidationError:
            self.fail("Empty valid Excel file was rejected")

    def test_rejects_doc_file(self):
        file = SimpleUploadedFile(
            "dummy.doc",
            b"fake word",
            content_type="application/msword",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_docx_file(self):
        file = SimpleUploadedFile(
            "dummy.docx",
            b"fake word",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_ppt_file(self):
        file = SimpleUploadedFile(
            "dummy.ppt",
            b"fake ppt",
            content_type="application/vnd.ms-powerpoint",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_pptx_file(self):
        file = SimpleUploadedFile(
            "dummy.pptx",
            b"fake ppt",
            content_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_csv_file(self):
        file = SimpleUploadedFile(
            "dummy.csv",
            b"a,b,c\n1,2,3",
            content_type="text/csv",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_image_file(self):
        file = SimpleUploadedFile(
            "image.jpg",
            b"\xff\xd8\xff\xe0",
            content_type="image/jpeg",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_png_file(self):
        file = SimpleUploadedFile(
            "image.png",
            b"\x89PNG\r\n\x1a\n",
            content_type="image/png",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_gif_file(self):
        file = SimpleUploadedFile(
            "image.gif",
            b"GIF89a",
            content_type="image/gif",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)

    def test_rejects_zip_file(self):
        file = SimpleUploadedFile(
            "archive.zip",
            b"PK\x03\x04",
            content_type="application/zip",
        )
        with self.assertRaises(ValidationError):
            validate_excel_file(file)


class ExcelSnifferTests(TestCase):
    def _temp_path(self, suffix):
        f = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        f.close()
        return f.name

    def test_valid_xlsx_file(self):
        file_path = self._temp_path(".xlsx")
        wb = Workbook()
        wb.save(file_path)

        sniffer = ExcelSniffer()
        self.assertIsNone(sniffer.is_valid(file_path))

    def test_valid_empty_xlsx_file(self):
        file_path = self._temp_path(".xlsx")
        wb = Workbook()
        wb.save(file_path)

        sniffer = ExcelSniffer()
        self.assertIsNone(sniffer.is_valid(file_path))

    def test_invalid_xlsx_file(self):
        file_path = self._temp_path(".xlsx")
        with open(file_path, "wb") as f:
            f.write(b"not really excel")

        sniffer = ExcelSniffer()
        with self.assertRaises(ValidationError):
            sniffer.is_valid(file_path)

    def test_fake_xlsx_with_wrong_content(self):
        file_path = self._temp_path(".xlsx")
        with open(file_path, "w") as f:
            f.write("Hello world")

        sniffer = ExcelSniffer()
        with self.assertRaises(ValidationError):
            sniffer.is_valid(file_path)

    def test_valid_xls_file(self):
        if not XlsWorkbook:
            self.skipTest("xlrd/xlwt not installed")

        file_path = self._temp_path(".xls")
        wb = XlsWorkbook()
        sheet = wb.add_sheet("Sheet1")
        sheet.write(0, 0, "Hello XLS")
        wb.save(file_path)

        sniffer = ExcelSniffer()
        self.assertIsNone(sniffer.is_valid(file_path))

    def test_invalid_xls_file(self):
        if not xlrd:
            self.skipTest("xlrd not installed")

        file_path = self._temp_path(".xls")
        with open(file_path, "wb") as f:
            f.write(b"not really an xls")

        sniffer = ExcelSniffer()
        with self.assertRaises(ValidationError):
            sniffer.is_valid(file_path)


class ExcelReaderTests(TestCase):
    def _xlsx_file(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append([1, "Pondasi", 10.5, "m3"])
        ws.append([2, "Beton Kolom", "1.000,5", "m3"])  # Indonesian number style
        wb.save(bio)
        return SimpleUploadedFile(
            "rab.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _xls_file(self):
        if not XlsWorkbook:
            self.skipTest("xlwt not installed")
        bio = BytesIO()
        wb = XlsWorkbook()
        ws = wb.add_sheet("Sheet1")
        rows = [
            ["No", "Uraian Pekerjaan", "Volume", "Satuan"],
            [1, "Pondasi", 3.25, "m3"],
            ["1.2", "Urugan", "2.500,00", "m3"],
        ]
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.write(r, c, val)
        wb.save(bio)
        return SimpleUploadedFile(
            "rab.xls",
            bio.getvalue(),
            content_type="application/vnd.ms-excel",
        )

    def test_read_xlsx_into_db(self):
        f = self._xlsx_file()
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 2)
        self.assertEqual(RabEntry.objects.count(), 2)

        first = RabEntry.objects.order_by("row_index").first()
        self.assertEqual(first.description, "Pondasi")
        self.assertEqual(first.unit, "m3")
        self.assertEqual(first.volume, Decimal("10.5"))

        # Indonesian-style number parsed
        second = RabEntry.objects.order_by("row_index")[1]
        self.assertEqual(second.volume, Decimal("1000.5"))

    def test_read_xls_into_db(self):
        f = self._xls_file()
        # test skipped above if xlwt not installed
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 2)
        self.assertEqual(RabEntry.objects.count(), 2)

    def test_header_variants(self):
        # Accept e.g. "Uraian", "Deskripsi", "Qty" for robustness
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian", "Qty", "Satuan"])
        ws.append(["A", "Pekerjaan A", "12", "m2"])
        wb.save(bio)
        f = SimpleUploadedFile(
            "alt.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 1)
        item = RabEntry.objects.get()
        self.assertEqual(item.item_number, "A")
        self.assertEqual(item.unit, "m2")
        self.assertEqual(item.volume, Decimal("12"))

    def test_reject_unsupported_type(self):
        bad = SimpleUploadedFile("data.csv", b"a,b,c\n", content_type="text/csv")
        with self.assertRaises(UnsupportedFileError):
            ExcelImporter().import_file(bad)


class HeaderMapperTests(TestCase):
    def test_maps_clean_headers(self):
        headers = ["No", "Uraian Pekerjaan", "Volume", "Satuan"]
        mapping, missing, originals = map_headers(headers)
        self.assertIn("no", mapping)
        self.assertIn("uraian", mapping)
        self.assertIn("volume", mapping)
        self.assertIn("satuan", mapping)
        self.assertEqual(missing, [])
        self.assertEqual(originals["uraian"], "Uraian Pekerjaan")

    def test_maps_with_punctuation_and_case(self):
        headers = ["No.", "URAIAN", "VOL.", "satuan "]
        mapping, missing, _ = map_headers(headers)
        self.assertTrue(set(["no", "uraian", "volume", "satuan"]).issubset(mapping.keys()))
        self.assertEqual(missing, [])

    def test_reports_missing_required(self):
        headers = ["No", "Deskripsi"]  # missing volume & satuan
        _, missing, _ = map_headers(headers)
        self.assertIn("volume", missing)
        self.assertIn("satuan", missing)

    def test_find_header_row_skips_title_block(self):
        rows = [
            ["PEKERJAAN", None, ":", "PEMBANGUNAN"],  # fake title
            [None, None, None],
            ["No.", "URAIAN PEKERJAAN", "VOL.", "SATUAN"],  # real header
            ["1", "Bata Merah", "10", "m"],
        ]
        idx = find_header_row(rows, scan_first=10)
        self.assertEqual(idx, 2)  # 0-based index


class RabParserTests(TestCase):
    def setUp(self):
        """Set up test dependencies and sample project"""
        self.parser = create_rab_parser()  # Uses factory with proper DI
        self.project = Project.objects.create(
            program="TEST PROGRAM",
            kegiatan="TEST ACTIVITY",
            pekerjaan="TEST JOB",
            lokasi="TEST LOCATION",
            tahun_anggaran=2025
        )

    def test_trims_leading_and_trailing_whitespace(self):
        self.assertEqual(self.parser.clean_cell("  Some Text  "), "Some Text")

    def test_collapses_multiple_internal_spaces(self):
        self.assertEqual(self.parser.clean_cell("Text   with    extra spaces"), "Text with extra spaces")

    def test_replaces_newline_with_space(self):
        self.assertEqual(self.parser.clean_cell("Line one\nLine two"), "Line one Line two")

    def test_empty_and_whitespace_cell_is_parsed_as_none(self):
        self.assertIsNone(self.parser.clean_cell(""))
        self.assertIsNone(self.parser.clean_cell("   "))

    def test_cell_with_non_breaking_space_is_parsed_as_none(self):
        self.assertIsNone(self.parser.clean_cell("\u00A0"))
        self.assertIsNone(self.parser.clean_cell(" \u00A0 "))

    def test_converts_standard_numeric_string_to_decimal(self):
        self.assertEqual(self.parser.to_decimal("1234.56"), Decimal("1234.56"))

    def test_converts_number_to_integer(self):
        self.assertEqual(self.parser.to_decimal("123"), Decimal("123"))

    def test_converts_currency_string_with_rp_to_decimal(self):
        self.assertEqual(self.parser.to_decimal("Rp 5.000"), Decimal("5000"))

    def test_converts_number_with_indonesian_format_to_decimal(self):
        self.assertEqual(self.parser.to_decimal("1.234.567,89"), Decimal("1234567.89"))

    def test_converts_us_format_with_commas_to_decimal(self):
        self.assertEqual(self.parser.to_decimal("1,234,567.89"), Decimal("1234567.89"))

    def test_converts_percentage_format_to_decimal(self):
        result = self.parser.to_percentage("75%")
        self.assertEqual(result, Decimal("0.75"))

    def test_converts_boolean_strings_true(self):
        self.assertTrue(self.parser.to_boolean("TRUE"))
        self.assertTrue(self.parser.to_boolean("true"))
        self.assertTrue(self.parser.to_boolean("True"))

    def test_converts_boolean_strings_false(self):
        self.assertFalse(self.parser.to_boolean("FALSE"))
        self.assertFalse(self.parser.to_boolean("false"))
        self.assertFalse(self.parser.to_boolean("False"))

    def test_converts_date_string(self):
        result = self.parser.to_date("2024-12-31")
        self.assertEqual(result, date(2024, 12, 31))

    def test_converts_indonesian_date_format(self):
        result = self.parser.to_date("31/12/2024")
        self.assertEqual(result, date(2024, 12, 31))

    def test_handles_large_numbers_without_crashing(self):
        self.assertEqual(self.parser.to_decimal("99.999.999.999,00"), Decimal("99999999999.00"))

    def test_handles_overflow_numbers_safely(self):
        result = self.parser.to_decimal("999999999999999999999999999999")
        self.assertIsInstance(result, Decimal)

    def test_handles_scientific_notation_string(self):
        self.assertEqual(self.parser.to_decimal("1.23E+5"), Decimal("123000"))

    def test_returns_none_for_invalid_numeric_string(self):
        self.assertIsNone(self.parser.to_decimal("12abc"))
        self.assertIsNone(self.parser.to_decimal("N/A"))

    def test_returns_none_for_empty_or_none_input(self):
        self.assertIsNone(self.parser.to_decimal(None))
        self.assertIsNone(self.parser.to_decimal(""))

    def test_handles_mixed_data_in_numeric_column(self):
        self.assertEqual(self.parser.clean_cell("N/A"), "N/A")
        self.assertIsNone(self.parser.to_decimal("N/A"))

    def test_parses_clean_row_and_creates_item_entry(self):
        raw_row = {
            'No.': '1', 'URAIAN PEKERJAAN': 'Mobilisasi/demobilisasi', 'SATUAN': 'Ls',
            'KODE ANALISA': 'AT.19-1', 'VOL.': '1,00', 'HARGA SATUAN (Rp.)': '1.500.000',
            'JUMLAH HARGA (Rp.)': '1.500.000'
        }
        parent_section = RabEntry.objects.create(
            project=self.project, entry_type=RabEntry.EntryType.SECTION, description="Parent"
        )
        entry = self.parser.parse_row(raw_row, project=self.project, parent=parent_section)

        self.assertEqual(entry.entry_type, RabEntry.EntryType.ITEM)
        self.assertEqual(entry.item_number, '1')
        self.assertEqual(entry.description, 'Mobilisasi/demobilisasi')
        self.assertEqual(entry.volume, Decimal('1.00'))
        self.assertEqual(entry.total_price, Decimal('1500000'))
        self.assertEqual(entry.parent, parent_section)

    def test_row_with_missing_middle_cell_is_parsed_with_null(self):
        raw_row = {'No.': '2', 'URAIAN PEKERJAAN': 'Test Item', 'VOL.': '5.00', 'SATUAN': None}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertIsNone(entry.unit)
        self.assertEqual(entry.volume, Decimal('5.00'))

    def test_row_with_only_whitespace_is_skipped(self):
        raw_row = {'No.': '   ', 'URAIAN PEKERJAAN': ' ', 'VOL.': None}
        self.assertIsNone(self.parser.parse_row(raw_row, project=self.project))

    def test_parses_row_with_extra_columns_gracefully(self):
        raw_row = {'No.': '3', 'URAIAN PEKERJAAN': 'Extra Item', 'EXTRA_COLUMN': 'ignored'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.description, 'Extra Item')

    def test_row_with_fewer_columns_pads_with_null(self):
        raw_row = {'No.': '4', 'URAIAN PEKERJAAN': 'Incomplete Item'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertIsNone(entry.volume)
        self.assertIsNone(entry.unit)

    def test_row_with_mixed_data_types_parsed_correctly(self):
        raw_row = {
            'No.': '5', 'URAIAN PEKERJAAN': 'Mixed Data Item',
            'VOL.': '2.50', 'HARGA SATUAN (Rp.)': '1.000.000,50'
        }
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.item_number, '5')
        self.assertEqual(entry.volume, Decimal('2.50'))
        self.assertEqual(entry.unit_price, Decimal('1000000.50'))

    def test_validates_required_fields(self):
        raw_row = {'No.': '6', 'URAIAN PEKERJAAN': '', 'VOL.': '1.00'}
        result = self.parser.parse_row(raw_row, project=self.project)
        self.assertIsNone(result)

    def test_handles_file_with_only_headers(self):
        initial_count = RabEntry.objects.count()
        row_list = []
        for row in row_list:
            self.parser.parse_row(row, self.project)
        self.assertEqual(RabEntry.objects.count(), initial_count)

    def test_processes_all_cells_for_data_type_conversion(self):
        raw_row = {
            'No.': ' 1 ', 'URAIAN PEKERJAAN': '  Test\nItem  ', 'SATUAN': ' Ls ',
            'KODE ANALISA': '  AT.19-1  ', 'VOL.': ' 2,50 ',
            'HARGA SATUAN (Rp.)': ' Rp 1.000.000 ', 'JUMLAH HARGA (Rp.)': ' 2.500.000,00 '
        }
        entry = self.parser.parse_row(raw_row, project=self.project)

        self.assertEqual(entry.item_number, '1')  # Trimmed
        self.assertEqual(entry.description, 'Test Item')  # Newline replaced, trimmed
        self.assertEqual(entry.unit, 'Ls')  # Trimmed
        self.assertEqual(entry.analysis_code, 'AT.19-1')  # Trimmed
        self.assertEqual(entry.volume, Decimal('2.50'))  # Indonesian format converted
        self.assertEqual(entry.unit_price, Decimal('1000000'))  # Currency converted
        self.assertEqual(entry.total_price, Decimal('2500000.00'))  # Indonesian format converted

    def test_preserves_na_values_for_manual_input(self):
        raw_row = {
            'No.': '1', 'URAIAN PEKERJAAN': 'N/A requires manual input',
            'SATUAN': 'N/A', 'KODE ANALISA': 'TBD'
        }
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.unit, 'N/A')
        self.assertEqual(entry.analysis_code, 'TBD')

    def test_classifies_row_as_site_header(self):
        raw_row = {'URAIAN PEKERJAAN': 'BUJAK 1', 'No.': None}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.entry_type, RabEntry.EntryType.SITE_HEADER)
        self.assertEqual(entry.description, "BUJAK 1")

    def test_classifies_row_as_main_section(self):
        raw_row = {'No.': 'A', 'URAIAN PEKERJAAN': 'PEMBUATAN SUMUR BOR'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.entry_type, RabEntry.EntryType.SECTION)

    def test_classifies_row_as_sub_section(self):
        raw_row = {'No.': 'I', 'URAIAN PEKERJAAN': 'PEKERJAAN PERSIAPAN'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.entry_type, RabEntry.EntryType.SUB_SECTION)

    def test_classifies_row_as_sub_total(self):
        raw_row = {'URAIAN PEKERJAAN': 'Sub Total Pekerjaan Persiapan', 'JUMLAH HARGA (Rp.)': '5.000.000'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.entry_type, RabEntry.EntryType.SUB_TOTAL)
        self.assertEqual(entry.total_price, Decimal('5000000'))

    def test_classifies_row_as_grand_total(self):
        raw_row = {'URAIAN PEKERJAAN': 'Total Biaya', 'JUMLAH HARGA (Rp.)': '125.000.000'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.entry_type, RabEntry.EntryType.SUB_TOTAL)

    def test_parses_row_with_total_in_words(self):
        raw_row = {'URAIAN PEKERJAAN': 'Terbilang : Seratus Juta Rupiah'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.entry_type, RabEntry.EntryType.GRAND_TOTAL)
        self.assertEqual(entry.total_price_in_words, "Seratus Juta Rupiah")

    def test_parses_comment_or_descriptive_row_as_item(self):
        raw_row = {'URAIAN PEKERJAAN': 'Pekerjaan dilaksanakan Sesuai Spek'}
        entry = self.parser.parse_row(raw_row, project=self.project)
        self.assertEqual(entry.entry_type, RabEntry.EntryType.ITEM)
        self.assertIsNone(entry.volume)


@override_settings(DATABASES=SQLITE_DB_SETTINGS)
class PreviewRowsOverrideTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.parser = create_rab_parser()
        self.project = Project.objects.create(
            program="TEST PROGRAM",
            kegiatan="TEST ACTIVITY",
            pekerjaan="TEST JOB",
            lokasi="TEST LOCATION",
            tahun_anggaran=2025
        )

    @patch("excel_parser.views.preview_file")
    @patch("excel_parser.views.validate_excel_file")
    def test_session_overrides_applied_to_preview(self, mock_validate, mock_preview):
        mock_validate.return_value = None
        base_row = {
            "row_key": "0000-override",
            "number": "1",
            "description": "Pekerjaan Sample",
            "volume": "1.00",
            "unit": "m2",
            "analysis_code": "AA-01",
            "price": "120.00",
            "total_price": "120.00",
            "is_section": False,
            "index_kind": None,
            "section_letter": None,
            "section_roman": None,
            "section_type": None,
            "job_match_status": "auto",
            "job_match": [],
            "job_match_error": None,
        }
        mock_preview.return_value = [base_row.copy()]

        session = self.client.session
        session["rab_overrides"] = {
            "0000-override": {
                "unit_price": "250.00",
                "total_price": "750.00",
                "volume": "3.00",
                "analysis_code": "ZZ-99",
            }
        }
        session.save()

        dummy = SimpleUploadedFile(
            "sample.xlsx",
            b"dummy",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post("/excel_parser/preview_rows", {"file": dummy})
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertIn("rows", payload)
        updated_row = payload["rows"][0]
        self.assertEqual(updated_row["price"], "250.00")
        self.assertEqual(updated_row["total_price"], "750.00")
        self.assertEqual(updated_row["volume"], "3.00")
        self.assertEqual(updated_row["analysis_code"], "ZZ-99")


class ReaderPrivateHelpersTests(TestCase):
    def test__norm_and__match_header(self):
        self.assertEqual(reader_mod._norm("  No.  "), "no.")
        key, raw = reader_mod._match_header("Uraian")
        self.assertEqual(key, "description")
        self.assertEqual(raw, "Uraian")
        key2, raw2 = reader_mod._match_header("Satuan")
        self.assertEqual(key2, "unit")
        self.assertEqual(raw2, "Satuan")

    def test_parse_decimal_full_branches(self):
        # invalid numeric strings should just return 0 instead of crashing
        self.assertEqual(reader_mod.parse_decimal("not_a_number"), Decimal("0"))
        self.assertEqual(reader_mod.parse_decimal("7 = 5 x 6"), Decimal("0"))

    def test_parse_decimal_branch_last_separator_logic(self):
        # Case 1: last comma after dot -> koma decimal
        self.assertEqual(reader_mod.parse_decimal("1.234,56"), Decimal("1234.56"))
        # Case 2: last dot after comma -> dot decimal
        self.assertEqual(reader_mod.parse_decimal("1,234.56"), Decimal("1234.56"))

    def test__ext_of_and_make_reader_and_unsupported(self):
        good_xlsx = SimpleUploadedFile("a.xlsx", b"data")
        self.assertEqual(reader_mod._ext_of(good_xlsx), "xlsx")
        r = reader_mod.make_reader(good_xlsx)
        self.assertIsInstance(r, reader_mod._XLSXReader)

        bad = SimpleUploadedFile("a.csv", b"a,b\n")
        self.assertEqual(reader_mod._ext_of(bad), "")
        with self.assertRaises(reader_mod.UnsupportedFileError):
            reader_mod.make_reader(bad)

    def test__base_reader_iter_rows_raises(self):
        base = reader_mod._BaseReader()
        with self.assertRaises(NotImplementedError):
            list(base.iter_rows(SimpleUploadedFile("x.xlsx", b"")))

    def test_parse_decimal_last_separator_fallback_irregular_grouping(self):
        self.assertEqual(reader_mod.parse_decimal("12.34,56"), Decimal("1234.56"))
        self.assertEqual(reader_mod.parse_decimal("1.2.3,4"), Decimal("123.4"))
        self.assertEqual(reader_mod.parse_decimal("12,34.56"), Decimal("1234.56"))
        self.assertEqual(reader_mod.parse_decimal("1,2,3.4"), Decimal("123.4"))


class ReaderHeaderAndRowParsingTests(TestCase):
    def test__find_header_map_success_with_preface(self):
        rows = [
            ["Judul", "Meta"],
            [None, None],
            ["No.", "Uraian Pekerjaan", "Volume", "Satuan"],
            ["1", "Item A", "1,5", "pcs"],
        ]
        colmap, hdr = reader_mod._find_header_map(rows)
        self.assertEqual(hdr, 2)
        for k in ("number", "description", "volume", "unit"):
            self.assertIn(k, colmap)

    def test__find_header_map_missing_raises(self):
        rows = [
            ["Random", "Header"],
            ["Kode", "Deskripsi"],
            ["1", "A"],
        ]
        with self.assertRaises(reader_mod.ParseError):
            reader_mod._find_header_map(rows)

    def test__rows_after(self):
        cache = [
            ["No", "Uraian", "Volume", "Satuan"],
            ["1", "A", "10", "m"],
            ["2", "B", "20", "m"],
        ]
        out = list(reader_mod._rows_after(cache, start_idx=0))
        self.assertEqual(len(out), 2)
        self.assertEqual(out[0][0], "1")
        self.assertEqual(out[1][1], "B")

    def test__parse_rows_skip_empty_and_cast(self):
        cache = [
            ["No", "Uraian", "Volume", "Satuan"],  # header index 0
            ["1", "Item A", "1.000,50", "m3"],
            ["", "", "", ""],  # empty -> skip
            ["2", "Item B", "2,5", "m3"],
        ]
        colmap = {"number": 0, "description": 1, "volume": 2, "unit": 3, "_header_row": 0}
        parsed = reader_mod._parse_rows(cache, colmap)
        self.assertEqual(len(parsed), 2)
        self.assertEqual(parsed[0].number, "1")
        self.assertEqual(parsed[0].description, "Item A")
        self.assertEqual(parsed[0].volume, Decimal("1000.50"))
        self.assertEqual(parsed[0].unit, "m3")
        self.assertEqual(parsed[1].volume, Decimal("2.5"))


class ImporterDefaultsTests(TestCase):
    def test_importer_sets_project_defaults_and_enum_and_count(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append(["1", "Gali tanah", "1.000,50", "m3"])
        ws.append(["2", "Urug tanah", "250", "m3"])
        wb.save(bio)
        f = SimpleUploadedFile(
            "mini.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        count = ExcelImporter().import_file(f)
        self.assertEqual(count, 2)

        proj = Project.objects.get()
        self.assertEqual(proj.program, "Default Program")
        self.assertEqual(proj.kegiatan, "Default Activity")
        self.assertEqual(proj.pekerjaan, "Imported from Excel")
        self.assertEqual(proj.lokasi, "Not Specified")
        self.assertEqual(proj.tahun_anggaran, 2025)

        self.assertEqual(RabEntry.objects.count(), 2)
        e1 = RabEntry.objects.order_by("row_index").first()
        self.assertEqual(e1.item_number, "1")
        self.assertEqual(e1.description, "Gali tanah")
        self.assertEqual(e1.volume, Decimal("1000.50"))
        self.assertEqual(e1.unit, "m3")
        self.assertEqual(e1.entry_type, RabEntry.EntryType.ITEM)


class PreviewFileTests(APITestCase):
    def test_preview_file_includes_all_columns(self):
        # Build Excel file in memory
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan", "Harga Satuan", "Jumlah Harga"])
        ws.append(["1", "Gali tanah", "1000", "m3", "25000", "25000000"])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        f = SimpleUploadedFile(
            "mini.xlsx",
            bio.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            "/excel_parser/preview_rows",
            {"excel_standard": f},
            format="multipart"
        )

        self.assertEqual(response.status_code, 200, response.content)
        rows = response.json()["excel_standard"]

        self.assertIn("analysis_code", rows[0])
        self.assertIn("price", rows[0])
        self.assertIn("total_price", rows[0])

    def test_preview_file_detects_section_rows(self):
        # Build an in-memory Excel file
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan", "Harga Satuan", "Jumlah Harga"])
        ws.append(["I.", "PEKERJAAN PERSIAPAN", "", ""])  # should be a section
        ws.append(["1", "Mobilisasi", "1", "Ls", "1000", "1000"])  # should be an item
        wb.save(bio)

        f = SimpleUploadedFile(
            "mini.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        rows = preview_file(f)
        # Check first row is section
        self.assertTrue(rows[0]["is_section"])
        self.assertEqual(rows[0]["description"], "PEKERJAAN PERSIAPAN")
        self.assertEqual(rows[0]["job_match_status"], "skipped")

        # Check second row is normal item
        self.assertFalse(rows[1]["is_section"])
        self.assertEqual(rows[1]["description"], "Mobilisasi")
        self.assertAlmostEqual(float(rows[1]["volume"]), 1.0, places=2)
        self.assertEqual(rows[1]["price"], "1000.00")
        self.assertEqual(rows[1]["total_price"], "1000.00")

    @patch("excel_parser.services.reader.match_description")
    def test_preview_file_includes_job_matching_results(self, mock_match):
        mock_match.return_value = {"status": "found", "match": {"code": "X.01", "name": "Dummy"}}

        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan", "Harga Satuan", "Jumlah Harga"])
        ws.append(["1", "Mobilisasi", "1", "Ls", "1000", "1000"])
        wb.save(bio)

        f = SimpleUploadedFile(
            "match.xlsx",
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        rows = preview_file(f)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["job_match_status"], "found")
        self.assertEqual(rows[0]["job_match"].get("code"), "X.01")
        mock_match.assert_called_once_with("Mobilisasi")


class TestProject(models.Model):
    name = models.CharField(max_length=255)

    class Meta:
        db_table = "test_projects"
        managed = True


class TestRab(models.Model):
    name = models.CharField(max_length=255)
    project = models.ForeignKey(TestProject, on_delete=models.CASCADE)

    class Meta:
        db_table = "test_rabs"
        managed = True


class TestRabItemHeader(models.Model):
    rab = models.ForeignKey(TestRab, on_delete=models.CASCADE)
    name = models.CharField(max_length=255)

    class Meta:
        db_table = "test_rab_item_headers"
        managed = True


class TestRabItem(models.Model):
    rab = models.ForeignKey(TestRab, on_delete=models.CASCADE, null=True)
    rab_item_header = models.ForeignKey(TestRabItemHeader, on_delete=models.CASCADE, null=True)
    name = models.CharField(max_length=500, null=True)
    volume = models.FloatField(null=True)
    price = models.FloatField(null=True)
    unit = models.CharField(max_length=50, null=True)

    class Meta:
        db_table = "test_rab_items"
        managed = True


class ImporterIntegrationTests(TransactionTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        # manually create test tables
        with connection.schema_editor() as schema:
            schema.create_model(TestProject)
            schema.create_model(TestRab)
            schema.create_model(TestRabItemHeader)
            schema.create_model(TestRabItem)

    @classmethod
    def tearDownClass(cls):
        # clean up tables
        with connection.schema_editor() as schema:
            schema.delete_model(TestRabItem)
            schema.delete_model(TestRabItemHeader)
            schema.delete_model(TestRab)
            schema.delete_model(TestProject)
        super().tearDownClass()

    def test_importer_inserts_into_rab_items(self):
        project = TestProject.objects.create(name="Test Project")
        rab = TestRab.objects.create(name="Test RAB", project=project)

        # build a tiny Excel file
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan", "Harga Satuan", "Jumlah Harga"])
        ws.append(["1", "Mobilisasi", "1", "Ls", "1000", "1000"])
        wb.save(bio)

        f = SimpleUploadedFile(
            "mini.xlsx", bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        rows = preview_file(f)

        # simulate writing into TestRabItem
        for row in rows:
            TestRabItem.objects.create(
                rab=rab,
                name=row["description"],
                volume=row["volume"],
                price=row["price"],
                unit=row["unit"]
            )

        self.assertEqual(TestRabItem.objects.count(), 1)
        item = TestRabItem.objects.first()
        self.assertEqual(item.name, "Mobilisasi")
        self.assertEqual(item.volume, 1.0)
        self.assertEqual(item.price, 1000.0)


class RABConvertedViewTests(TestCase):
    def setUp(self):
        self.client = Client()

    def test_rab_converted_page_loads(self):
        response = self.client.get("/excel_parser/rab_converted/")
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "rab_converted.html")


class LoggingTests(TestCase):
    def _make_excel_file(self, filename="logtest.xlsx"):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append([1, "Test pekerjaan", 10, "m2"])
        wb.save(bio)
        return SimpleUploadedFile(
            filename, bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_import_file_logs(self):
        f = self._make_excel_file()
        importer = ExcelImporter()

        with self.assertLogs("excel_parser", level="INFO") as cm:
            importer.import_file(f)

        logs = "\n".join(cm.output)
        self.assertIn("Starting import", logs)
        self.assertIn("Import finished", logs)

    def test_preview_file_logs(self):
        f = self._make_excel_file()

        with self.assertLogs("excel_parser", level="INFO") as cm:
            preview_file(f)

        logs = "\n".join(cm.output)
        self.assertIn("Previewing file", logs)
        self.assertIn("Preview parsed", logs)

    def test_validator_logs_warning(self):
        bad_file = SimpleUploadedFile("bad.txt", b"not excel", content_type="text/plain")

        with self.assertLogs("excel_parser", level="WARNING") as cm:
            with self.assertRaises(ValidationError):
                validate_excel_file(bad_file)

        logs = "\n".join(cm.output)
        self.assertIn("Rejected file", logs)


class DetectHeadersTests(APITestCase):
    def setUp(self):
        self.client = Client()

    def _make_workbook(self, header_row_index=2):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        # add some title rows to ensure header may be found lower
        for _ in range(header_row_index):
            ws.append([None, None, None])
        # actual header
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        # one data row
        ws.append([1, "Pondasi", 10, "m3"])
        wb.save(bio)
        bio.seek(0)
        return bio

    def test_detect_headers_success(self):
        bio = self._make_workbook(header_row_index=1)
        f = SimpleUploadedFile(
            "test.xlsx", bio.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp = self.client.post("/excel_parser/detect_headers", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        # header_row_index is 1-based in response
        self.assertEqual(payload["header_row_index"], 3)
        self.assertIn("mapping", payload)
        self.assertIn("originals", payload)
        self.assertIn("missing", payload)

    def test_detect_headers_missing_file(self):
        resp = self.client.post("/excel_parser/detect_headers", {}, format="multipart")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("file is required", resp.json().get("detail", ""))

    def test_detect_headers_invalid_excel(self):
        # send a pdf (invalid) so validate_excel_file should raise -> 400
        bad = SimpleUploadedFile("bad.pdf", b"%PDF", content_type="application/pdf")
        resp = self.client.post("/excel_parser/detect_headers", {"file": bad}, format="multipart")
        # service validator rejects non-excel -> 400
        self.assertEqual(resp.status_code, 400)

    @patch("excel_parser.views.find_header_row")
    def test_detect_headers_header_not_found(self, mock_find):
        # simulate find_header_row returning -1
        mock_find.return_value = -1
        bio = self._make_workbook()
        f = SimpleUploadedFile(
            "test.xlsx", bio.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp = self.client.post("/excel_parser/detect_headers", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, 422)
        self.assertIn("header not found", resp.json().get("detail", ""))

    @patch("excel_parser.views.load_workbook")
    def test_detect_headers_unexpected_exception(self, mock_load):
        mock_load.side_effect = Exception("boom")
        bio = self._make_workbook()
        f = SimpleUploadedFile(
            "test.xlsx", bio.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp = self.client.post("/excel_parser/detect_headers", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, 500)
        self.assertIn("boom", resp.json().get("detail", ""))


class PreviewRowsExtendedPathsTests(APITestCase):
    def setUp(self):
        self.client = Client()

    def _make_excel_bytes(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan", "Harga Satuan", "Jumlah Harga"])
        ws.append(["1", "Mobilisasi", "1", "Ls", "1000", "1000"])
        wb.save(bio)
        return bio.getvalue()

    def test_preview_rows_no_file(self):
        resp = self.client.post("/excel_parser/preview_rows", {}, format="multipart")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("No file uploaded", resp.json().get("detail", ""))

    def test_preview_rows_get_not_allowed(self):
        resp = self.client.get("/excel_parser/preview_rows")
        self.assertEqual(resp.status_code, 405)

    def test_preview_rows_pdf_accepts_and_reports(self):
        pdf = SimpleUploadedFile("dummy.pdf", b"%PDF", content_type="application/pdf")
        resp = self.client.post("/excel_parser/preview_rows", {"pdf_file": pdf}, format="multipart")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        self.assertIn("pdf_file", payload)
        self.assertIn("PDF uploaded successfully", payload["pdf_file"]["message"])

    def test_preview_rows_pdf_rejects_wrong_mimetype(self):
        # wrong mimetype should raise ValidationError from validate_pdf_file -> 400
        fake = SimpleUploadedFile("dummy.pdf", b"not pdf", content_type="text/plain")
        resp = self.client.post("/excel_parser/preview_rows", {"pdf_file": fake}, format="multipart")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("Only .pdf files are allowed", resp.json().get("error", ""))

    def test_preview_rows_excel_standard_and_apendo(self):
        b = self._make_excel_bytes()
        std = SimpleUploadedFile("std.xlsx", b, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        apendo = SimpleUploadedFile("apendo.xlsx", b, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp = self.client.post("/excel_parser/preview_rows", {"excel_standard": std, "excel_apendo": apendo}, format="multipart")
        self.assertEqual(resp.status_code, 200)
        payload = resp.json()
        # both keys should be present
        self.assertIn("excel_standard", payload)
        self.assertIn("excel_apendo", payload)
        # ensure returned rows include expected keys
        self.assertIn("description", payload["excel_standard"][0])
        self.assertIn("price", payload["excel_standard"][0])


class UploadViewTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _make_excel(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append([1, "Test pekerjaan", 10, "m2"])
        wb.save(bio)
        return bio.getvalue()

    def test_upload_view_get_renders(self):
        req = self.factory.get("/excel_parser/upload")
        resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 200)

    def test_upload_view_post_excel_standard_success(self):
        """Cover lines 138-139: if excel_standard: validate_excel_file(excel_standard)"""
        excel = SimpleUploadedFile(
            "up.xlsx", self._make_excel(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        req = self.factory.post("/excel_parser/upload", {}, FILES={"excel_standard": excel})
        resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Standard Excel uploaded successfully", resp.content.decode())

    def test_upload_view_post_apendo_success(self):
        """Cover lines 144-145: if excel_apendo: validate_excel_file(excel_apendo) return render"""
        excel = SimpleUploadedFile(
            "apendo.xlsx", self._make_excel(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        req = self.factory.post("/excel_parser/upload", {}, FILES={"excel_apendo": excel})
        resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("APENDO Excel uploaded successfully", resp.content.decode())

    def test_upload_view_post_pdf_success(self):
        """Cover lines 150-151: if pdf_file: validate_pdf_file(pdf_file) return render"""
        pdf = SimpleUploadedFile("file.pdf", b"%PDF", content_type="application/pdf")
        req = self.factory.post("/excel_parser/upload", {}, FILES={"pdf_file": pdf})
        resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("PDF uploaded successfully", resp.content.decode())

    def test_upload_view_no_file_returns_400(self):
        req = self.factory.post("/excel_parser/upload", {})
        resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 400)
        self.assertIn("No file selected", resp.content.decode())

    def test_upload_view_validation_error_returns_400(self):
        """Cover lines 159-161: except ValidationError as ve: return render with error"""
        fake = SimpleUploadedFile("file.pdf", b"notpdf", content_type="text/plain")
        req = self.factory.post("/excel_parser/upload", {}, FILES={"pdf_file": fake})
        resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 400)
        self.assertIn("Only .pdf files are allowed", resp.content.decode())

    def test_upload_view_generic_exception_returns_500(self):
        """Cover lines 162-165: except Exception as e: return render with error"""
        excel = SimpleUploadedFile(
            "test.xlsx", self._make_excel(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        req = self.factory.post("/excel_parser/upload", {}, FILES={"excel_standard": excel})

        with patch('excel_parser.views.validate_excel_file') as mock_validate:
            mock_validate.side_effect = Exception("Unexpected error")
            resp = views.upload_view(req)

        self.assertEqual(resp.status_code, 500)
        self.assertIn("Unexpected error", resp.content.decode())

class ApplyPreviewOverridesUnitTests(TestCase):
    def test_no_overrides_returns_early(self):
        rows = [{"row_key": "k1", "volume": "1"}]
        views._apply_preview_overrides(rows, None)
        self.assertEqual(rows[0]["volume"], "1")

    def test_row_without_row_key_is_ignored(self):
        rows = [{"description": "no key"}]
        overrides = {"k1": {"volume": "9"}}
        views._apply_preview_overrides(rows, overrides)
        self.assertEqual(rows[0]["description"], "no key")

    def test_row_with_no_override_data_is_ignored(self):
        rows = [{"row_key": "k-not-found", "volume": "1"}]
        overrides = {"other": {"volume": "9"}}
        views._apply_preview_overrides(rows, overrides)
        self.assertEqual(rows[0]["volume"], "1")

    def test_apply_all_override_fields_and_price_fallback(self):
        rows = [{"row_key": "rk1", "volume": "1", "price": "100.00", "total_price": "100.00"}]
        overrides = {"rk1": {"volume": "2.5", "analysis_code": "AC-1", "unit_price": "250.00", "total_price": "625.00"}}
        views._apply_preview_overrides(rows, overrides)
        r = rows[0]
        self.assertEqual(r["volume"], "2.5")
        self.assertEqual(r["analysis_code"], "AC-1")
        self.assertEqual(r["price"], "250.00")
        self.assertEqual(r["total_price"], "625.00")

    def test_price_fallback_uses_price_key_when_unit_price_missing(self):
        rows = [{"row_key": "rk2", "price": "99.00", "total_price": "99.00"}]
        overrides = {"rk2": {"price": "199.00", "total_price": "199.00"}}
        views._apply_preview_overrides(rows, overrides)
        self.assertEqual(rows[0]["price"], "199.00")
        self.assertEqual(rows[0]["total_price"], "199.00")

    def test_apply_preview_overrides_with_partial_keys(self):
        rows = [
            {"row_key": "k1", "volume": "1", "price": "10.00"},
            {"row_key": "k2", "volume": "2", "price": "20.00"},
            {"description": "no key row"},
        ]
        overrides = {
            "k1": {"volume": "5.5"},
            "k2": {"price": "99.99"},
        }
        views._apply_preview_overrides(rows, overrides)
        self.assertEqual(rows[0]["volume"], "5.5")
        self.assertEqual(rows[0]["price"], "10.00")
        self.assertEqual(rows[1]["price"], "99.99")
        self.assertEqual(rows[2]["description"], "no key row")

    def test_apply_preview_overrides_total_price_only_and_volume_only(self):
        rows = [
            {"row_key": "r1", "price": "10.00", "total_price": "10.00", "volume": "1"},
            {"row_key": "r2", "price": "5.00", "total_price": "5.00", "volume": "1"},
        ]
        overrides = {
            "r1": {"total_price": "99.99"},
            "r2": {"volume": "3.5"},
        }
        views._apply_preview_overrides(rows, overrides)
        self.assertEqual(rows[0]["total_price"], "99.99")
        self.assertEqual(rows[0]["price"], "10.00")
        self.assertEqual(rows[1]["volume"], "3.5")
        self.assertEqual(rows[1]["price"], "5.00")


class PreviewRowsExceptionPathsTests(TestCase):
    def setUp(self):
        self.client = Client()

    def _make_excel_bytes(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append(["1", "Mobilisasi", "1", "Ls"])
        wb.save(bio)
        return bio.getvalue()

    def test_preview_rows_raises_validation_error_for_excel(self):
        bad = SimpleUploadedFile("bad.xlsx", b"not excel", content_type="text/plain")
        resp = self.client.post("/excel_parser/preview_rows", {"excel_standard": bad}, format="multipart")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("error", resp.json())

    @patch("excel_parser.views.preview_file")
    def test_preview_rows_unexpected_exception_returns_500(self, mock_preview):
        mock_preview.side_effect = Exception("boom-preview")
        good = SimpleUploadedFile(
            "good.xlsx",
            self._make_excel_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp = self.client.post("/excel_parser/preview_rows", {"excel_standard": good}, format="multipart")
        self.assertEqual(resp.status_code, 500)
        self.assertIn("boom-preview", resp.json().get("error", ""))


class PreviewRowsExtendedOverridesAndErrors(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_extended_path_applies_session_overrides_to_both_keys(self):
        row = {
            "row_key": "OVR-1",
            "number": "1",
            "description": "Pekerjaan A",
            "volume": "1.00",
            "unit": "m2",
            "analysis_code": None,
            "price": "1000.00",
            "total_price": "1000.00",
        }

        with patch("excel_parser.views.preview_file", return_value=[row.copy()]) as mock_preview:
            payload = {
                "excel_standard": SimpleUploadedFile(
                    "std.xlsx", _make_xlsx_bytes(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                "excel_apendo": SimpleUploadedFile(
                    "apendo.xlsx", _make_xlsx_bytes(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            }

            req = self.factory.post("/excel_parser/preview_rows", {}, FILES=payload)
            _attach_session(req)
            req.session["rab_overrides"] = {
                "OVR-1": {
                    "unit_price": "250.00",
                    "total_price": "500.00",
                    "volume": "2.00",
                    "analysis_code": "AC-TEST",
                }
            }
            req.session.save()

            resp = views.preview_rows(req)
            self.assertEqual(resp.status_code, 200)
            data = resp.json()
            self.assertIn("excel_standard", data)
            self.assertIn("excel_apendo", data)
            std_row = data["excel_standard"][0]
            apendo_row = data["excel_apendo"][0]

            for r in (std_row, apendo_row):
                self.assertEqual(r["price"], "250.00")
                self.assertEqual(r["total_price"], "500.00")
                self.assertEqual(r["volume"], "2.00")
                self.assertEqual(r["analysis_code"], "AC-TEST")

            assert mock_preview.call_count == 2

    def test_extended_path_validationerror_from_validator_returns_400(self):
        f = SimpleUploadedFile(
            "apendo.xlsx", _make_xlsx_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        req = self.factory.post("/excel_parser/preview_rows", {}, FILES={"excel_apendo": f})
        _attach_session(req)

        with patch("excel_parser.views.validate_excel_file", side_effect=ValidationError("bad-excel")):
            resp = views.preview_rows(req)
        self.assertEqual(resp.status_code, 400)
        self.assertIn("bad-excel", resp.json().get("error", ""))


class UploadViewBranchesTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _make_sample_xlsx(self):
        bio = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.append(["No", "Uraian Pekerjaan", "Volume", "Satuan"])
        ws.append([1, "Test pekerjaan", 10, "m2"])
        wb.save(bio)
        return bio.getvalue()

    @patch("excel_parser.views.validate_excel_file")
    def test_upload_view_excel_standard_success_calls_validator_and_renders(self, mock_validate):
        mock_validate.return_value = None
        excel = SimpleUploadedFile(
            "std.xlsx", self._make_sample_xlsx(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        req = self.factory.post("/excel_parser/upload", {}, FILES={"excel_standard": excel})
        resp = views.upload_view(req)
        mock_validate.assert_called_once()
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Standard Excel uploaded successfully", resp.content.decode())

    @patch("excel_parser.views.validate_excel_file")
    def test_upload_view_excel_apendo_success_calls_validator_and_renders(self, mock_validate):
        mock_validate.return_value = None
        excel = SimpleUploadedFile(
            "apendo.xlsx", self._make_sample_xlsx(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        req = self.factory.post("/excel_parser/upload", {}, FILES={"excel_apendo": excel})
        resp = views.upload_view(req)
        mock_validate.assert_called_once()
        self.assertEqual(resp.status_code, 200)
        self.assertIn("APENDO Excel uploaded successfully", resp.content.decode())

    @patch("excel_parser.views.validate_pdf_file")
    def test_upload_view_pdf_success_calls_validator_and_renders(self, mock_validate_pdf):
        mock_validate_pdf.return_value = None
        pdf = SimpleUploadedFile("file.pdf", b"%PDF", content_type="application/pdf")
        req = self.factory.post("/excel_parser/upload", {}, FILES={"pdf_file": pdf})
        resp = views.upload_view(req)
        mock_validate_pdf.assert_called_once()
        self.assertEqual(resp.status_code, 200)
        self.assertIn("PDF uploaded successfully", resp.content.decode())

    def test_upload_view_validate_excel_raises_validationerror_renders_400(self):
        excel = SimpleUploadedFile(
            "bad.xlsx", self._make_sample_xlsx(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        req = self.factory.post("/excel_parser/upload", {}, FILES={"excel_standard": excel})
        with patch("excel_parser.views.validate_excel_file", side_effect=ValidationError("bad-excel")):
            resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 400)
        self.assertIn("bad-excel", resp.content.decode())

    def test_upload_view_validate_pdf_raises_unexpected_exception_renders_500(self):
        pdf = SimpleUploadedFile("file.pdf", b"%PDF", content_type="application/pdf")
        req = self.factory.post("/excel_parser/upload", {}, FILES={"pdf_file": pdf})
        with patch("excel_parser.views.validate_pdf_file", side_effect=Exception("boom-pdf")):
            resp = views.upload_view(req)
        self.assertEqual(resp.status_code, 500)
        self.assertIn("boom-pdf", resp.content.decode())

    def test_validate_pdf_file_accepts_and_rejects(self):
        good = SimpleUploadedFile("ok.pdf", b"%PDF-1.4\n", content_type="application/pdf")
        views.validate_pdf_file(good)

        bad = SimpleUploadedFile("bad.pdf", b"notpdf", content_type="text/plain")
        with self.assertRaises(ValidationError):
            views.validate_pdf_file(bad)

class ViewsTestCase(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_preview_rows_pdf_file(self):
        pdf_content = b'%PDF-1.4 test pdf'
        pdf_file = SimpleUploadedFile("test.pdf", pdf_content, content_type="application/pdf")
        request = self.factory.post('/excel_parser/preview_rows', {'pdf_file': pdf_file})
        request.FILES['pdf_file'] = pdf_file
        request.session = {}
        response = views.preview_rows(request)
        self.assertEqual(response.status_code, 200)
        self.assertIn("pdf_file", response.json())

    def test_upload_view_excel_standard(self):
        excel_content = b'PK\x03\x04'  # minimal XLSX header
        excel_file = SimpleUploadedFile("test.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        request = self.factory.post('/excel_parser/upload', {'excel_standard': excel_file})
        request.FILES['excel_standard'] = excel_file
        response = views.upload_view(request)
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'Standard Excel uploaded successfully', response.content)

    def test_upload_view_excel_apendo(self):
        excel_content = b'PK\x03\x04'
        excel_file = SimpleUploadedFile("apendo.xlsx", excel_content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        request = self.factory.post('/excel_parser/upload', {'excel_apendo': excel_file})
        request.FILES['excel_apendo'] = excel_file
        response = views.upload_view(request)
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'APENDO Excel uploaded successfully', response.content)

    def test_upload_view_pdf_file(self):
        pdf_content = b'%PDF-1.4 test pdf'
        pdf_file = SimpleUploadedFile("test.pdf", pdf_content, content_type="application/pdf")
        request = self.factory.post('/excel_parser/upload', {'pdf_file': pdf_file})
        request.FILES['pdf_file'] = pdf_file
        response = views.upload_view(request)
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'PDF uploaded successfully', response.content)

    def test_upload_view_validation_error(self):
        # Pass a file with wrong content type to trigger ValidationError
        bad_file = SimpleUploadedFile("bad.txt", b"bad", content_type="text/plain")
        request = self.factory.post('/excel_parser/upload', {'pdf_file': bad_file})
        request.FILES['pdf_file'] = bad_file
        response = views.upload_view(request)
        self.assertEqual(response.status_code, 400)
        self.assertIn(b'Only .pdf files are allowed.', response.content)

    def test_upload_view_exception(self):
        # Patch validate_pdf_file to raise an Exception
        def raise_exception(_):
            raise Exception("Test exception")
        orig = views.validate_pdf_file
        views.validate_pdf_file = raise_exception
        pdf_content = b'%PDF-1.4 test pdf'
        pdf_file = SimpleUploadedFile("test.pdf", pdf_content, content_type="application/pdf")
        request = self.factory.post('/excel_parser/upload', {'pdf_file': pdf_file})
        request.FILES['pdf_file'] = pdf_file
        response = views.upload_view(request)
        self.assertEqual(response.status_code, 500)
        self.assertIn(b'Test exception', response.content)
        views.validate_pdf_file = orig

class MinimalDirectTests(TestCase):
    """Minimal tests using RequestFactory but calling upload_view directly"""
    
    def test_upload_pdf_direct_call(self):
        """Direct function call to upload_view with PDF"""
        factory = RequestFactory()
        pdf_file = SimpleUploadedFile("t.pdf", b"%PDF-1.4", content_type="application/pdf")
        
        request = factory.post('/upload', {})
        request.FILES['pdf_file'] = pdf_file
        
        response = views.upload_view(request)
        
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'PDF uploaded successfully', response.content)
    
    def test_upload_invalid_pdf_direct_call(self):
        """Direct function call to upload_view with invalid PDF"""
        factory = RequestFactory()
        bad_file = SimpleUploadedFile("bad.pdf", b"not pdf", content_type="image/png")
        
        request = factory.post('/upload', {})
        request.FILES['pdf_file'] = bad_file
        
        response = views.upload_view(request)
        
        self.assertEqual(response.status_code, 400)
        self.assertIn(b'Only .pdf files are allowed', response.content)