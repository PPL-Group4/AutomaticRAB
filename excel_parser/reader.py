from __future__ import annotations
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import re
from typing import List, Dict, Iterable, Tuple

from django.core.files.uploadedfile import UploadedFile

from excel_parser.models import RABItem

class UnsupportedFileError(Exception):
    pass

class ParseError(Exception):
    pass

HEADER_ALIASES = {
    "number": {"no", "no.", "nomor", "number", "kode"},
    "description": {"uraian pekerjaan", "uraian", "deskripsi", "pekerjaan", "job description"},
    "volume": {"volume", "qty", "jumlah", "kuantitas"},
    "unit": {"satuan", "unit"},
}

def _norm(s) -> str:
    return str(s or "").strip().lower()

def _match_header(cell: str) -> Tuple[str | None, str]:
    """Return (canonical_key, raw) or (None, raw)."""
    key = None
    n = _norm(cell)
    for canon, aliases in HEADER_ALIASES.items():
        if n in aliases:
            key = canon
            break
    return key, cell

_THOUSAND_DOT_DECIMAL_COMMA = re.compile(r"^\d{1,3}(\.\d{3})+,\d+$")
_THOUSAND_COMMA_DECIMAL_DOT = re.compile(r"^\d{1,3}(,\d{3})+\.\d+$")

def parse_decimal(val) -> Decimal:
    if val is None or str(val).strip() == "":
        return Decimal("0")
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))

    s = str(val).strip()
    # 1.000,50 -> 1000.50
    if _THOUSAND_DOT_DECIMAL_COMMA.match(s):
        s = s.replace(".", "").replace(",", ".")
    # 1,000.50 -> 1000.50
    elif _THOUSAND_COMMA_DECIMAL_DOT.match(s):
        s = s.replace(",", "")
    else:
        # if it ends with ",xx" assume comma-decimal
        if "," in s and "." not in s:
            s = s.replace(".", "").replace(",", ".")
        # if it has both, guess last separator is decimal
        elif "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")

    try:
        return Decimal(s)
    except InvalidOperation:
        raise ParseError(f"Cannot parse decimal from '{val}'")

class _BaseReader:
    def iter_rows(self, file: UploadedFile) -> Iterable[List]:
        raise NotImplementedError

class _XLSXReader(_BaseReader):
    def iter_rows(self, file: UploadedFile) -> Iterable[List]:
        from openpyxl import load_workbook
        pos = file.tell()
        file.seek(0)
        wb = load_workbook(filename=file, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        file.seek(pos)

class _XLSReader(_BaseReader):
    def iter_rows(self, file: UploadedFile) -> Iterable[List]:
        import xlrd
        pos = file.tell()
        data = file.read()
        wb = xlrd.open_workbook(file_contents=data)
        sh = wb.sheet_by_index(0)
        for r in range(sh.nrows):
            yield [sh.cell_value(r, c) for c in range(sh.ncols)]
        file.seek(pos)

def _ext_of(file: UploadedFile) -> str:
    name = (file.name or "").lower()
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".xls"):
        return "xls"
    return ""

def make_reader(file: UploadedFile) -> _BaseReader:
    ext = _ext_of(file)
    if ext == "xlsx":
        return _XLSXReader()
    if ext == "xls":
        return _XLSReader()
    raise UnsupportedFileError("Only .xls and .xlsx are supported")

@dataclass
class ParsedRow:
    number: str
    description: str
    volume: Decimal
    unit: str

def _find_header_map(rows: Iterable[List]) -> Tuple[Dict[str, int], int]:
    """
    Scan up to first 10 rows to detect header line and map columns.
    Returns (mapping, header_row_index)
    """
    cache = list(rows)
    limit = min(len(cache), 10)
    for i in range(limit):
        row = cache[i]
        seen = {}
        for idx, cell in enumerate(row):
            canon, _ = _match_header(cell)
            if canon and canon not in seen:
                seen[canon] = idx
        if {"number", "description", "volume", "unit"} <= set(seen.keys()):
            return seen, i
    raise ParseError("Required headers not found (No, Uraian Pekerjaan, Volume, Satuan)")

def _rows_after(cache: List[List], start_idx: int) -> Iterable[List]:
    for r in range(start_idx + 1, len(cache)):
        yield cache[r]

def _parse_rows(cache: List[List], colmap: Dict[str, int]) -> List[ParsedRow]:
    out: List[ParsedRow] = []
    for idx, row in enumerate(_rows_after(cache, start_idx=colmap["_header_row"])):  # type: ignore
        def cell(col):
            i = colmap[col]
            return row[i] if i < len(row) else None

        desc = (cell("description") or "").strip() if isinstance(cell("description"), str) else cell("description")
        if not desc:   # stop on empty rows
            continue

        number = str(cell("number") or "").strip()
        unit = str(cell("unit") or "").strip()
        volume = parse_decimal(cell("volume"))

        out.append(ParsedRow(number=number, description=str(desc), volume=volume, unit=unit))
    return out

class ExcelImporter:
    """
    High-level façade used by views/tasks/tests.
    SRP: import an uploaded excel into RABItem rows
    OCP: new readers can be added without modifying this class
    """
    def import_file(self, file: UploadedFile) -> int:
        reader = make_reader(file)
        # materialize all rows to detect header first
        cache = list(reader.iter_rows(file))
        colmap, header_row = _find_header_map(cache)
        colmap["_header_row"] = header_row  # keep header position

        parsed = _parse_rows(cache, colmap)
        count = 0
        for idx, p in enumerate(parsed, start=1):
            RABItem.objects.create(
                number=p.number,
                description=p.description,
                volume=p.volume,
                unit=p.unit,
                source_filename=file.name,
                row_index=idx,
            )
            count += 1
        return count
