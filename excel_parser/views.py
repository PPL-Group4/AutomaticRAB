from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
from django.views import View
from django.shortcuts import render, redirect

from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response

from openpyxl import load_workbook

from .services.header_mapper import map_headers, find_header_row
from .services.reader import preview_file
from .services.validators import validate_excel_file

# SENTRY +: import sentry_sdk untuk performance monitoring
import sentry_sdk


def validate_pdf_file(file):
    """Special validator for PDF files."""
    if file.content_type != "application/pdf":
        raise ValidationError("Only .pdf files are allowed.")


@api_view(['POST'])
@parser_classes([MultiPartParser])
def detect_headers(request):
    """
    POST /excel_parser/detect_headers
    body: multipart form-data with 'file'
    """
    f = request.FILES.get('file')
    if not f:
        return Response({"detail": "file is required"}, status=400)

    # SENTRY +: kasih tag supaya bisa difilter di Sentry
    with sentry_sdk.configure_scope() as scope:
        scope.set_tag("feature", "excel_detect_headers")
        scope.set_tag("excel.file_name", getattr(f, "name", ""))
        scope.set_tag("excel.file_size_bytes", getattr(f, "size", 0))
        scope.set_tag("excel.content_type", getattr(f, "content_type", ""))

    try:
        # SENTRY +: span untuk validasi file
        with sentry_sdk.start_span(
            op="excel.validate",
            description="Validate Excel file for header detection"
        ):
            validate_excel_file(f)

        # SENTRY +: span untuk load workbook & baca 200 baris pertama
        with sentry_sdk.start_span(
            op="excel.load_workbook",
            description="Load workbook for header detection"
        ):
            wb = load_workbook(f, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [list(r) for r in ws.iter_rows(values_only=True, min_row=1, max_row=200)]

        # SENTRY +: span untuk cari header & mapping
        with sentry_sdk.start_span(
            op="excel.detect_header",
            description="Find header row and map headers"
        ):
            hdr_idx = find_header_row(rows)
            if hdr_idx < 0:
                return Response({"detail": "header not found"}, status=422)

            header_row = [str(c or '') for c in rows[hdr_idx]]
            mapping, missing, originals = map_headers(header_row)

        return Response({
            "sheet": ws.title,
            "header_row_index": hdr_idx + 1,  # 1-based index
            "mapping": mapping,
            "originals": originals,
            "missing": missing
        })

    except ValidationError as ve:
        return Response({"detail": str(ve)}, status=400)


@csrf_exempt
def preview_rows(request):
    """
    POST /excel_parser/preview_rows
    Accepts either:
      - file (legacy, for frontend JS)
      - excel_standard / excel_apendo / pdf_file (extended, from git)
    """
    if request.method != "POST":
        return JsonResponse({"detail": "Method not allowed"}, status=405)

    # legacy (frontend expects 'file')
    legacy_file = request.FILES.get("file")

    # extended (multiple file support)
    excel_standard = request.FILES.get("excel_standard")
    excel_apendo = request.FILES.get("excel_apendo")
    pdf_file = request.FILES.get("pdf_file")

    # SENTRY +: tag umum untuk endpoint preview
    with sentry_sdk.configure_scope() as scope:
        scope.set_tag("feature", "excel_preview")
        scope.set_tag("excel.has_legacy_file", bool(legacy_file))
        scope.set_tag("excel.has_standard_file", bool(excel_standard))
        scope.set_tag("excel.has_apendo_file", bool(excel_apendo))
        scope.set_tag("excel.has_pdf_file", bool(pdf_file))

    try:
        results = {}

        # === Legacy path ===
        if legacy_file:
            # SENTRY +: span khusus untuk preview legacy
            with sentry_sdk.start_span(
                op="excel.preview.legacy",
                description="Preview legacy Excel file"
            ):
                validate_excel_file(legacy_file)
                results["rows"] = preview_file(legacy_file)

        # === Extended path ===
        if excel_standard:
            # SENTRY +: span untuk preview standard
            with sentry_sdk.start_span(
                op="excel.preview.standard",
                description="Preview standard Excel file"
            ):
                validate_excel_file(excel_standard)
                results["excel_standard"] = preview_file(excel_standard)

        if excel_apendo:
            # SENTRY +: span untuk preview APENDO
            with sentry_sdk.start_span(
                op="excel.preview.apendo",
                description="Preview APENDO Excel file"
            ):
                validate_excel_file(excel_apendo)
                results["excel_apendo"] = preview_file(excel_apendo)

        if pdf_file:
            # SENTRY +: span untuk upload PDF (belum ada parser)
            with sentry_sdk.start_span(
                op="excel.preview.pdf",
                description="Upload PDF file (no parser yet)"
            ):
                validate_pdf_file(pdf_file)
                results["pdf_file"] = {"message": "PDF uploaded successfully (no parser yet)"}

        if not results:
            return JsonResponse({"detail": "No file uploaded"}, status=400)

        return JsonResponse(results, safe=False)

    except ValidationError as ve:
        return JsonResponse({"error": str(ve)}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def upload_view(request):
    """
    Render upload page for Excel & PDF with format validation.
    """
    if request.method == 'POST':
        excel_standard = request.FILES.get("excel_standard")
        excel_apendo = request.FILES.get("excel_apendo")
        pdf_file = request.FILES.get("pdf_file")

        try:
            if excel_standard:
                validate_excel_file(excel_standard)
                return render(request, 'excel_upload.html', {
                    'success': 'Standard Excel uploaded successfully'
                })

            if excel_apendo:
                validate_excel_file(excel_apendo)
                TEMPLATE_UPLOAD = 'excel_upload.html'
                return render(request, TEMPLATE_UPLOAD, {
                    'success': 'APENDO Excel uploaded successfully'
                })

            if pdf_file:
                validate_pdf_file(pdf_file)
                return render(request, TEMPLATE_UPLOAD, {
                    'success': 'PDF uploaded successfully'
                })

            return render(request, TEMPLATE_UPLOAD, {
                'error': 'No file selected'
            }, status=400)

        except ValidationError as ve:
            return render(request, TEMPLATE_UPLOAD, {
                'error': str(ve)
            }, status=400)

    return render(request, TEMPLATE_UPLOAD)


def rab_converted(request):
    """
    Display converted rows in a table with ability to edit.
    This will hit the preview_rows endpoint via AJAX/fetch.
    """
    return render(request, "rab_converted.html")
