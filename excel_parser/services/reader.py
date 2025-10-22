from __future__ import annotations
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import hashlib
import re
from typing import List, Dict, Iterable, Tuple, Optional
import string

from django.core.files.uploadedfile import UploadedFile
from excel_parser.models import Project, RabEntry
from .job_matcher import match_description

import logging
logger = logging.getLogger("excel_parser")

class UnsupportedFileError(Exception):
    pass

class ParseError(Exception):
    pass

HEADER_ALIASES = {
    "number": {"no", "no.", "nomor", "number", "kode"},
    "description": {"uraian pekerjaan", "uraian", "deskripsi", "pekerjaan", "job description"},
    "volume": {"volume", "vol", "vol.", "qty", "jumlah", "kuantitas"},
    "unit": {"satuan", "unit"},
    "analysis_code": {"kode analisa", "kode", "analysis code"},
    "price": {"harga satuan", "harga_satuan", "price", "harga"},
    "total_price": {"jumlah harga", "total harga", "total", "total_price"}
}

def _norm(s) -> str:
    return str(s or "").strip().lower()

_THOUSAND_DOT_DECIMAL_COMMA = re.compile(r"^\d{1,3}(\.\d{3})+,\d+$")
_THOUSAND_COMMA_DECIMAL_DOT = re.compile(r"^\d{1,3}(,\d{3})+\.\d+$")

def classify_index_token(token: str) -> str:
    """Classify a No. cell into 'letter' | 'roman' | 'numeric' | 'none'."""
    if token is None:
        return "none"
    t = str(token).strip()
    if not t:
        return "none"
    t = t.rstrip(" .)")
    up = t.upper()
    if len(up) == 1 and up in string.ascii_uppercase:
        return "letter"
    if re.fullmatch(r"[IVXLCDM]+", up):
        return "roman"
    tn = up.replace(".", "").replace(",", "")
    if tn.isdigit():
        return "numeric"
    return "none"

def parse_decimal(val) -> Decimal:
    if val is None or str(val).strip() == "":
        return Decimal("0")
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))

    s = str(val).strip()
    if not any(ch.isdigit() for ch in s):
        return Decimal("0")
    if "=" in s or re.search(r"\d+\s*[xX]\s*\d+", s):
        return Decimal("0")
    if _THOUSAND_DOT_DECIMAL_COMMA.match(s):
        s = s.replace(".", "").replace(",", ".")
    elif _THOUSAND_COMMA_DECIMAL_DOT.match(s):
        s = s.replace(",", "")
    else:
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        elif "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
    s = re.sub(r"[^\d.-]", "", s)
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")

class _BaseReader:
    def iter_rows(self, file: UploadedFile) -> Iterable[List]:
        raise NotImplementedError

class _XLSXReader(_BaseReader):
    def iter_rows(self, file: UploadedFile) -> Iterable[List]:
        from openpyxl import load_workbook
        pos = file.tell()
        file.seek(0)
        wb = load_workbook(filename=file, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
        file.seek(pos)

class _XLSReader(_BaseReader):
    def iter_rows(self, file: UploadedFile) -> Iterable[List]:
        import xlrd
        pos = file.tell()
        data = file.read()
        wb = xlrd.open_workbook(file_contents=data)
        sh = wb.sheet_by_index(0)
        for r in range(sh.nrows):
            yield [sh.cell_value(r, c) for c in range(sh.ncols)]
        file.seek(pos)

def _ext_of(file: UploadedFile) -> str:
    name = (file.name or "").lower()
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".xls"):
        return "xls"
    return ""

def make_reader(file: UploadedFile) -> _BaseReader:
    ext = _ext_of(file)
    if ext == "xlsx":
        return _XLSXReader()
    if ext == "xls":
        return _XLSReader()
    raise UnsupportedFileError("Only .xls and .xlsx are supported")

@dataclass
class ParsedRow:
    number: str
    description: str
    volume: Decimal
    unit: str
    analysis_code: str
    price: Decimal
    total_price: Decimal
    is_section: bool = False
    index_kind: str = "none"
    section_letter: str | None = None
    section_roman: str | None = None
    section_type: str | None = None  # NEW

def _find_header_map(rows: Iterable[List]) -> Tuple[Dict[str, int], int]:
    cache = list(rows)
    limit = min(len(cache), 50)
    for i in range(limit):
        row = cache[i]
        seen = {}
        for idx, cell in enumerate(row):
            if cell is None:
                continue
            normed = _norm(cell)
            canon = None
            for key, aliases in HEADER_ALIASES.items():
                if normed in {a.lower().strip() for a in aliases}:
                    canon = key
                    break
            if canon and canon not in seen:
                seen[canon] = idx
        if {"number", "description", "unit"} <= set(seen.keys()):
            return seen, i
    raise ParseError("Required headers not found (need at least No, Uraian Pekerjaan, and Satuan)")

def _rows_after(cache: List[List], start_idx: int) -> Iterable[List]:
    for r in range(start_idx + 1, len(cache)):
        yield cache[r]
def _match_header(cell: str) -> Tuple[str | None, str]:
    """Return (canonical_key, raw) or (None, raw)."""
    key = None
    n = _norm(cell)
    for canon, aliases in HEADER_ALIASES.items():
        if n in {a.lower().strip() for a in aliases}:
            key = canon
            break
    return key, cell


def _is_section_row(number: str, desc: str) -> bool:
    if number:
        kind = classify_index_token(number)
        if kind in ("letter", "roman"):
            return True
        if kind == "numeric":
            return False
    if desc and isinstance(desc, str) and desc.strip() and desc.strip().isupper():
        return True
    return False

def _parse_rows(cache: List[List], colmap: Dict[str, int]) -> List[ParsedRow]:
    out: List[ParsedRow] = []
    current_letter: str | None = None
    current_roman: str | None = None

    for idx, row in enumerate(_rows_after(cache, start_idx=colmap["_header_row"])):
        def cell(col):
            i = colmap.get(col)
            return row[i] if i is not None and i < len(row) else None

        number = str(cell("number") or "").strip()
        desc = (cell("description") or "").strip() if isinstance(cell("description"), str) else cell("description")
        if not desc:
            continue

        unit_val = str(cell("unit") or "").strip()
        vol_cell = cell("volume")

        index_kind = classify_index_token(number)
        is_section = _is_section_row(number, str(desc))

        if is_section:
            if index_kind == "letter":
                volume_result = parse_decimal(vol_cell)
            else:
                volume_result = Decimal("0")
        else:
            volume_result = parse_decimal(vol_cell)

        if is_section:
            if index_kind == "letter":
                current_letter = str(desc)
                current_roman = None
            elif index_kind == "roman":
                current_roman = str(desc)

        parsed_row = ParsedRow(
            number=number,
            description=str(desc),
            volume=volume_result,
            unit=unit_val,
            analysis_code=str(cell("analysis_code") or "").strip(),
            price=parse_decimal(cell("price")),
            total_price=parse_decimal(cell("total_price")),
            is_section=is_section,
            index_kind=index_kind,
            section_letter=current_letter,
            section_roman=current_roman,
            section_type="CATEGORY" if index_kind == "letter" else ("SECTION" if index_kind == "roman" else None)
        )
        out.append(parsed_row)
    return out

class ExcelImporter:
    def import_file(self, file: UploadedFile) -> int:
        logger.info("Starting import: file=%s", getattr(file, "name", "?"))

        reader = make_reader(file)
        cache = list(reader.iter_rows(file))
        logger.debug("Read %d rows from file=%s", len(cache), getattr(file, "name", "?"))

        colmap, header_row = _find_header_map(cache)
        logger.debug("Header row detected at index=%d, mapped columns=%s",
                     header_row, list(colmap.keys()))
        colmap["_header_row"] = header_row

        project, created = Project.objects.get_or_create(
            program="Default Program",
            kegiatan="Default Activity",
            pekerjaan="Imported from Excel",
            lokasi="Not Specified",
            tahun_anggaran=2025,
            defaults={'source_filename': file.name}
        )
        if created:
            logger.info("Created new default Project id=%s for source=%s", project.id, file.name)

        parsed = _parse_rows(cache, colmap)
        logger.info("Parsed %d rows from %s", len(parsed), file.name)

        count = 0
        for idx, p in enumerate(parsed, start=1):
            try:
                RabEntry.objects.create(
                    project=project,
                    entry_type=RabEntry.EntryType.SECTION if p.is_section else RabEntry.EntryType.ITEM,
                    item_number=p.number,
                    description=p.description,
                    volume=p.volume,
                    unit=p.unit,
                    analysis_code=p.analysis_code,
                    unit_price=p.price,
                    total_price=p.total_price,
                    row_index=idx,
                )
                count += 1
            except Exception:
                logger.exception("Failed to insert row %d into DB (file=%s)", idx, file.name)

        logger.info("Import finished: inserted=%d rows from %s", count, file.name)
        return count


def preview_file(file: UploadedFile):
    logger.info("Previewing file=%s", getattr(file, "name", "?"))

    reader = make_reader(file)
    cache = list(reader.iter_rows(file))
    logger.debug("Read %d rows during preview for file=%s", len(cache), getattr(file, "name", "?"))

    colmap, header_row = _find_header_map(cache)
    logger.debug("Preview header row index=%d, columns=%s", header_row, list(colmap.keys()))
    colmap["_header_row"] = header_row

    parsed = _parse_rows(cache, colmap)
    logger.info("Preview parsed %d rows from %s", len(parsed), file.name)

    from decimal import ROUND_HALF_UP

    preview_rows = []
    for idx, row in enumerate(parsed):
        if row.is_section:
            match_info = {"status": "skipped", "match": None}
        else:
            match_info = match_description(row.description)
            if not isinstance(match_info, dict):  # defensive guard for unexpected returns
                match_info = {"status": "error", "match": None, "error": "Unexpected match result"}

        row_key = _build_preview_row_key(row.description, row.number, idx)
        preview_rows.append(
            {
                "row_key": row_key,
                "number": row.number,
                "description": row.description,
                "volume": str(row.volume.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)),
                "unit": row.unit,
                "analysis_code": row.analysis_code,
                "price": str(row.price.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)),
                "total_price": str(row.total_price.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)),
                "is_section": row.is_section,
                "index_kind": row.index_kind,
                "section_letter": row.section_letter,
                "section_roman": row.section_roman,
                "section_type": row.section_type,
                "job_match_status": match_info.get("status"),
                "job_match": match_info.get("match"),
                "job_match_error": match_info.get("error"),
            }
        )

    return preview_rows


def _build_preview_row_key(description: Optional[str], number: Optional[str], index: int) -> str:
    """Create a stable key for a preview row used for override persistence."""
    normalized_desc = (description or "").strip().lower()
    normalized_num = (number or "").strip().lower()
    digest_source = f"{normalized_desc}|{normalized_num}".encode("utf-8", "ignore")
    digest = hashlib.sha1(digest_source).hexdigest()[:12]
    return f"{index:04d}-{digest}"
